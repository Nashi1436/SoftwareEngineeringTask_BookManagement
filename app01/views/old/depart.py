from django.shortcuts import render, redirect, HttpResponse
from app01 import models

from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm
from app01.utils.bootstrap import BootStrapModelForm

class UpModelForm(BootStrapModelForm):
    bootstrap_exclude_fields = ['img']

    class Meta:
        model = models.Department
        fields = "__all__"



def depart_list(request):
    """ 部门列表 """
    
    # 去数据库中获取所有的部门列表
    #  [对象,对象,对象]
    # queryset = models.Department.objects.all()

    # return render(request, 'depart_list.html', {'queryset': queryset})


    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["name__contains"] = search_data

    # 根据搜索条件去数据库获取
    queryset = models.Department.objects.filter(**data_dict).order_by('-id')
    # 分页
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        "search_data": search_data
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    """ 添加部门 """
    title = "添加部门"
    if request.method == "GET":
        form = UpModelForm()
        return render(request, 'change.html', {"form": form, 'title': title})

    form = UpModelForm(data=request.POST, files=request.FILES)
    if form.is_valid():
        form.save()
        return redirect("/depart/list/")
    return render(request, 'change.html', {"form": form, 'title': title})


def depart_edit(request, nid):
    """ 修改部门 """
     # 对象 / None
    row_object = models.Department.objects.filter(id=nid).first()
    if not row_object:
        # return render(request, 'error.html', {"msg": "数据不存在"})
        return redirect('/depart/list/')

    title = "编辑采集信息"
    if request.method == "GET":
        form = UpModelForm(instance=row_object)
        return render(request, 'change.html', {"form": form, "title": title})

    form = UpModelForm(data=request.POST, files=request.FILES, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/depart/list/')
    return render(request, 'change.html', {"form": form, "title": title})


def depart_delete(request, nid):
    """ 删除部门 """
    models.Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")






def depart_multi(request):
    """ 批量删除（Excel文件）"""
    from openpyxl import load_workbook

    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/depart/list/')
